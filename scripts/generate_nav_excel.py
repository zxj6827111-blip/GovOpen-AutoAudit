#!/usr/bin/env python3
"""
生成深度导航链接Excel报告
"""
import json
import os
import sys

# 确保openpyxl已安装
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Installing openpyxl...")
    os.system('pip install openpyxl -q')
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_navigation_report(batch_id, site_id="site_suqian_gov_zwgk"):
    """生成导航链接Excel报告"""
    
    trace_path = f"runs/{batch_id}/{site_id}/trace.json"
    
    if not os.path.exists(trace_path):
        print(f"错误: 找不到 {trace_path}")
        return
    
    # 读取trace数据
    with open(trace_path, 'r', encoding='utf-8') as f:
        trace = json.load(f)
    
    print(f"读取到 {len(trace)} 条导航记录")
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "深度导航链接记录"
    
    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头
    headers = ["序号", "导航类型", "栏目名称", "链接地址", "状态码", "截图文件"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # 栏目名称映射
    def extract_page_name(url):
        if not url:
            return ""
        path = url.split("/")[-1].replace(".shtml", "").replace(".html", "")
        
        # 常见栏目映射
        name_map = {
            "zwgk": "政务公开首页",
            "xxgk": "信息公开",
            "gkzn": "公开指南",
            "gkzd": "公开制度",
            "zdgk": "重点公开",
            "xxgkml": "信息公开目录",
            "jggk": "机构概况",
            "zcgzk": "政策规章库",
            "dfxfg": "地方性法规",
            "zfwjjd": "政府文件解读",
            "gknb": "公开年报",
            "cbzl": "财编资料",
        }
        
        for key, name in name_map.items():
            if key in url.lower():
                return name
        
        return path if path else "首页"
    
    # 填充数据
    for idx, item in enumerate(trace, 1):
        row = idx + 1
        url = item.get("url", "")
        step = item.get("step", "")
        status = item.get("status_code", "")
        
        step_name = "入口页" if step == "entry" else "深度导航"
        page_name = extract_page_name(url)
        screenshot = f"screenshot_{idx-1}.jpg"
        
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=step_name).border = border
        ws.cell(row=row, column=3, value=page_name).border = border
        ws.cell(row=row, column=4, value=url).border = border
        ws.cell(row=row, column=5, value=status).border = border
        ws.cell(row=row, column=6, value=screenshot).border = border
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 70
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 18
    
    # 保存
    output_path = f"runs/{batch_id}/深度导航链接记录.xlsx"
    wb.save(output_path)
    
    abs_path = os.path.abspath(output_path)
    print(f"\n✅ Excel报告已生成!")
    print(f"📁 文件路径: {abs_path}")
    print(f"📊 共记录 {len(trace)} 个链接")
    
    # 打印链接预览
    print("\n=== 链接预览 ===")
    for idx, item in enumerate(trace, 1):
        url = item.get("url", "")
        step = item.get("step", "")
        print(f"  {idx}. [{step}] {url}")
    
    return abs_path

if __name__ == "__main__":
    batch_id = sys.argv[1] if len(sys.argv) > 1 else "batch_060a64ab"
    generate_navigation_report(batch_id)
